import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
import io
import datetime

st.set_page_config(page_title="ระบบสรุปยอดค่าเทอม", layout="wide")
st.title("📊 โปรแกรมสรุปยอดค่าเทอมค้างชำระ (ดึงข้อมูลจาก Express)")

def process_csv(file):
    outstanding_data = []
    content = file.getvalue().decode('cp874', errors='ignore')
    reader = csv.reader(io.StringIO(content))
    
    for row in reader:
        if not row:
            continue
        non_empty = [x.strip() for x in row if x.strip()]
        
        if non_empty and non_empty[0] == "รวมลูกค้า":
            try:
                student_name = non_empty[1]
                student_code = non_empty[2]
                total_str = non_empty[-1].replace(',', '')
                outstanding_data.append({
                    'รหัสนักเรียน': student_code,
                    'ชื่อ-นามสกุล': student_name,
                    'ยอดค้างชำระ (บาท)': float(total_str),
                    'อ้างอิงจากไฟล์': file.name
                })
            except Exception:
                pass
    return outstanding_data

uploaded_files = st.file_uploader(
    "📂 ลากไฟล์ CSV จากโปรแกรม Express มาวางที่นี่ (เลือกได้หลายไฟล์พร้อมกัน)", 
    type=['csv'], 
    accept_multiple_files=True
)

if uploaded_files:
    if st.button("ประมวลผลข้อมูล"):
        all_data = []
        for file in uploaded_files:
            all_data.extend(process_csv(file))
            
        if all_data:
            df = pd.DataFrame(all_data)
            st.success(f"ดึงข้อมูลสำเร็จ! พบรายชื่อลูกหนี้ทั้งหมด {len(df)} รายการ")
            st.dataframe(df, use_container_width=True)
            
            # สร้างไฟล์ Excel ในหน่วยความจำ
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "รายงานยอดค้างชำระ"
            
            headers = ["ลำดับ", "รหัสนักเรียน", "ชื่อ-นามสกุล", "ยอดค้างชำระ (บาท)", "อ้างอิงจากไฟล์"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
            total_sum = 0
            for idx, r in df.iterrows():
                row_num = idx + 1
                s_code = r['รหัสนักเรียน']
                s_name = r['ชื่อ-นามสกุล']
                s_amount = r['ยอดค้างชำระ (บาท)']
                s_file = r['อ้างอิงจากไฟล์']
                
                ws.append([row_num, s_code, s_name, s_amount, s_file])
                total_sum += s_amount
                
            for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
                for cell in row:
                    cell.border = thin_border
                row[0].alignment = Alignment(horizontal='center')
                row[1].alignment = Alignment(horizontal='center')
                row[3].number_format = '#,##0.00'
                
            total_row = len(df) + 2
            ws.append(["", "", "รวมยอดค้างชำระทั้งหมด", total_sum, ""])
            for cell in ws[total_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = thin_border
            ws.cell(row=total_row, column=4).number_format = '#,##0.00'
            
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            
            wb.save(output)
            
            st.download_button(
                label="📥 ดาวน์โหลดไฟล์ Excel สรุปยอด",
                data=output.getvalue(),
                file_name=f"รายงานสรุปยอดค่าเทอม_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("ไม่พบข้อมูลลูกหนี้ในไฟล์ที่อัปโหลด กรุณาตรวจสอบไฟล์อีกครั้ง")
